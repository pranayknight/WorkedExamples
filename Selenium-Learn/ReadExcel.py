import openpyxl

path = "/home/pranay/Documents/data3.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active    # We can also use sheet=workbook.get_sheet_by_name("Sheet1")

row = sheet.max_row
col = sheet.max_column

print(row)
print(col)

for r in range(1, row+1):
    for c in range(1, col+1):
        print(sheet.cell(row=r, column=c).value, end="        ")

    print()